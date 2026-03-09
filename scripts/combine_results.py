#!/usr/bin/env python3
"""
Combine the benchmark dataset with model prediction results into a single JSON file
and an Excel spreadsheet for easy review.

Merges:
  - Dataset fields (problem_statement, patch, test_patch, FAIL_TO_PASS, etc.)
  - Agent prediction (model_patch, model_name)
  - Agent trajectory info (exit_status, cost, api_calls)
  - Eval results (resolved, exit_code, patch_successfully_applied, test_output)

Usage:
    python scripts/combine_results.py \
        --dataset internal-swe-bench-data/results_v1_gpt_5_2_68_20260307_verified.json \
        --preds internal-swe-bench-data/results/anthropic_claude_opus_4_6/preds.json \
        [--eval-dir eval_output/opus_eval/anthropic__claude-opus-4.6/] \
        [--output combined.json]

Output is saved next to preds.json as combined.json + combined.xlsx.
"""

import argparse
import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def load_dataset(path: str) -> dict:
    """Load dataset JSON and index by instance_id."""
    with open(path) as f:
        data = json.load(f)
    return {item["instance_id"]: item for item in data}


def load_preds(path: str) -> dict:
    """Load preds.json (dict keyed by instance_id)."""
    with open(path) as f:
        return json.load(f)


def load_trajectory(results_dir: str, instance_id: str) -> dict | None:
    """Load trajectory info from per-instance directory if it exists."""
    traj_dir = os.path.join(results_dir, instance_id)
    traj_file = os.path.join(traj_dir, f"{instance_id}.traj.json")
    if not os.path.isfile(traj_file):
        return None
    with open(traj_file) as f:
        traj = json.load(f)
    info = traj.get("info", {})
    stats = info.get("model_stats", {})
    return {
        "exit_status": info.get("exit_status"),
        "submission": info.get("submission"),
        "api_calls": stats.get("api_calls"),
        "instance_cost": stats.get("instance_cost"),
        "tokens_sent": stats.get("tokens_sent"),
        "tokens_received": stats.get("tokens_received"),
    }


def load_eval(eval_dir: str, instance_id: str) -> dict | None:
    """Load eval results (report.json + test_output.txt) for an instance."""
    inst_dir = os.path.join(eval_dir, instance_id)
    if not os.path.isdir(inst_dir):
        return None

    result = {}

    report_file = os.path.join(inst_dir, "report.json")
    if os.path.isfile(report_file):
        with open(report_file) as f:
            report = json.load(f)
        # report is {instance_id: {resolved, exit_code, ...}}
        result.update(report.get(instance_id, {}))

    test_output_file = os.path.join(inst_dir, "test_output.txt")
    if os.path.isfile(test_output_file):
        with open(test_output_file) as f:
            result["test_output"] = f.read()

    return result if result else None


def _cell_text(value) -> str:
    """Convert a value to display string for Excel."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, indent=2)
    s = str(value)
    # Ensure escaped \n becomes real newlines for display
    s = s.replace("\\n", "\n")
    return s


def _auto_col_width(values: list[str], min_w: int = 8, max_w: int = 80) -> float:
    """Calculate column width based on content. Uses the longest single line across all values."""
    longest = 0
    for v in values:
        for line in v.split("\n"):
            longest = max(longest, len(line))
    # Add a small padding; cap between min and max
    return min(max(longest + 2, min_w), max_w)


def _row_height(cells_text: list[str], col_widths: list[float], base_height: float = 15.0) -> float:
    """Estimate row height based on number of wrapped lines across all cells."""
    max_lines = 1
    for text, width in zip(cells_text, col_widths):
        if not text:
            continue
        lines = 0
        for line in text.split("\n"):
            # Each line wraps based on column width (approx 1 char = 1 width unit)
            char_width = max(width - 2, 1)
            lines += max(1, -(-len(line) // int(char_width)))  # ceil division
        max_lines = max(max_lines, lines)
    # Cap row height to avoid extremely tall rows
    max_lines = min(max_lines, 80)
    return max_lines * base_height


def export_excel(combined: list, xlsx_path: str):
    """Export combined data to an Excel file with auto-fit column widths and row heights."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SWE-bench Results"

    columns = [
        "instance_id", "repo", "model_name", "resolved",
        "exit_status", "eval_exit_code", "patch_successfully_applied",
        "api_calls", "instance_cost", "tokens_sent", "tokens_received",
        "problem_statement", "gold_patch", "model_patch", "test_patch",
        "FAIL_TO_PASS", "PASS_TO_PASS",
        "base_commit", "pull_number", "created_at", "version",
        "hints_text", "test_output",
    ]

    # Max width caps per column type
    long_text_cols = {"problem_statement", "gold_patch", "model_patch", "test_patch",
                      "hints_text", "test_output"}
    medium_text_cols = {"FAIL_TO_PASS", "PASS_TO_PASS"}

    # Pre-convert all cell values to strings
    all_text = []  # [row_idx][col_idx] = str
    for entry in combined:
        row = [_cell_text(entry.get(col)) for col in columns]
        all_text.append(row)

    # Header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    code_font = Font(name="Consolas", size=10)
    code_cols = {"gold_patch", "model_patch", "test_patch", "test_output"}
    for row_idx, (entry, row_text) in enumerate(zip(combined, all_text), 2):
        for col_idx, (col_name, text) in enumerate(zip(columns, row_text), 1):
            value = entry.get(col_name)
            # Keep numeric types as-is for Excel (not string)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
            elif isinstance(value, bool):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = wrap_alignment
            if col_name in code_cols:
                cell.font = code_font

    # Auto-fit column widths
    col_widths = []
    for col_idx, col_name in enumerate(columns):
        col_values = [col_name] + [row[col_idx] for row in all_text]
        if col_name in long_text_cols:
            max_w = 80
        elif col_name in medium_text_cols:
            max_w = 50
        else:
            max_w = 35
        w = _auto_col_width(col_values, min_w=8, max_w=max_w)
        col_widths.append(w)
        letter = ws.cell(row=1, column=col_idx + 1).column_letter
        ws.column_dimensions[letter].width = w

    # Auto-fit row heights
    ws.row_dimensions[1].height = 20  # header
    for row_idx, row_text in enumerate(all_text, 2):
        h = _row_height(row_text, col_widths)
        ws.row_dimensions[row_idx].height = h

    # Freeze header row
    ws.freeze_panes = "A2"

    # Highlight resolved=True rows in green
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for row_idx, entry in enumerate(combined, 2):
        if entry.get("resolved") is True:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = green_fill

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Combine dataset + predictions + eval results")
    parser.add_argument("--dataset", required=True, help="Path to the benchmark dataset JSON")
    parser.add_argument("--preds", required=True, help="Path to preds.json")
    parser.add_argument("--eval-dir", default=None, help="Path to eval output directory (optional)")
    parser.add_argument("--output", default=None, help="Output path (default: combined.json next to preds.json)")
    parser.add_argument("--model-name", default=None, help="Model name to tag in output (auto-detected from preds if not set)")
    args = parser.parse_args()

    dataset = load_dataset(args.dataset)
    preds = load_preds(args.preds)
    results_dir = str(Path(args.preds).parent)

    # Determine model name: from arg, or auto-detect from first pred entry
    model_name = args.model_name
    if not model_name:
        for pred in preds.values():
            if pred and pred.get("model_name_or_path"):
                model_name = pred["model_name_or_path"]
                break

    # Sanitize model name for filename
    model_name_safe = (model_name or "unknown").replace("/", "_").replace(" ", "_")

    if args.output:
        output_path = args.output
    else:
        output_path = os.path.join(results_dir, f"combined_{model_name_safe}.json")

    combined = []
    # Union of all instance IDs from dataset
    all_ids = sorted(dataset.keys())

    for instance_id in all_ids:
        entry = {}

        # Dataset fields
        ds = dataset[instance_id]
        entry["instance_id"] = instance_id
        entry["repo"] = ds.get("repo")
        entry["pull_number"] = ds.get("pull_number")
        entry["base_commit"] = ds.get("base_commit")
        entry["problem_statement"] = ds.get("problem_statement")
        entry["hints_text"] = ds.get("hints_text")
        entry["gold_patch"] = ds.get("patch")
        entry["test_patch"] = ds.get("test_patch")
        entry["FAIL_TO_PASS"] = ds.get("FAIL_TO_PASS")
        entry["PASS_TO_PASS"] = ds.get("PASS_TO_PASS")
        entry["created_at"] = ds.get("created_at")
        entry["version"] = ds.get("version")

        # Agent prediction
        pred = preds.get(instance_id)
        entry["model_name"] = model_name
        entry["model_patch"] = pred.get("model_patch") if pred else None

        # Trajectory info
        traj = load_trajectory(results_dir, instance_id)
        if traj:
            entry["exit_status"] = traj["exit_status"]
            entry["api_calls"] = traj["api_calls"]
            entry["instance_cost"] = traj["instance_cost"]
            entry["tokens_sent"] = traj["tokens_sent"]
            entry["tokens_received"] = traj["tokens_received"]

        # Eval results
        if args.eval_dir:
            eval_result = load_eval(args.eval_dir, instance_id)
            if eval_result:
                entry["resolved"] = eval_result.get("resolved")
                entry["eval_exit_code"] = eval_result.get("exit_code")
                entry["patch_successfully_applied"] = eval_result.get("patch_successfully_applied")
                entry["test_output"] = eval_result.get("test_output")

        combined.append(entry)

    with open(output_path, "w") as f:
        json.dump(combined, f, indent=2, ensure_ascii=False)

    # Export Excel
    xlsx_path = output_path.replace(".json", ".xlsx")
    export_excel(combined, xlsx_path)

    # Print summary
    total = len(combined)
    has_pred = sum(1 for e in combined if e.get("model_patch"))
    has_eval = sum(1 for e in combined if e.get("resolved") is not None)
    resolved = sum(1 for e in combined if e.get("resolved") is True)

    print(f"Combined {total} instances -> {output_path}")
    print(f"  Excel:       {xlsx_path}")
    print(f"  Predictions: {has_pred}/{total}")
    if has_eval:
        print(f"  Evaluated:   {has_eval}/{total}")
        print(f"  Resolved:    {resolved}/{has_eval}")


if __name__ == "__main__":
    main()
